import pandas as pd
import os
from .models import Fichier_mansuelle , Périmètre
from django.http import HttpResponseBadRequest, HttpResponseServerError, HttpResponse
import logging
from django.shortcuts import render, redirect
from reportlab.lib.pagesizes import A4, landscape
from .models import Utilisateur, Fichier_mansuelle ,Périmètre,Region,Prévision_perimetre,Visualisation
from django.db.models.functions import Coalesce
from io import BytesIO
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from django.db.models import Sum, F, FloatField,Value
from django.db.models import Sum, F, Value, FloatField
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from django.http import JsonResponse
from django.db.models import Count
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from .models import Périmètre, Fichier_mansuelle
from django.conf import settings
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Sum, Count, Q
from django.db.models import F, Value, Case, When
from django.db.models.functions import Coalesce
from django.db.utils import IntegrityError

def acceuil(request):
   return render(request,'page_acceuil.html')
def taritemnt_mansuel(request):
   return render(request,'taritemnt_mansuel.html')
def fichier_verifier(request):
   return render(request,'page_resultat_verifier.html')
def fichier_non_verifier(request):
   return render(request,'page_resultat_non_verifier.html')
def loginn(request):
   return render(request,'login.html')
def essay8(request):
   return render(request,'page_sauvgarde.html')
def traitement_p1(request):
   return render(request,'traitement_p1.html')
def traitement_annuel(request):
   return render(request,'traitement_annuel.html')
def dashboard(request):
   return render(request,'dashboard.html')
def page_reponce(request):
   return render(request,'page_réponce.html')

 
def login(request):
    nom_utilisateur_invalid = False
    mot_de_passe_invalid = False

    if request.method == 'POST':
        nom = request.POST.get('nom')
        password = request.POST.get('password')

        try:
            utilisateur = Utilisateur.objects.get(nom=nom)
            if utilisateur.password == password:
                request.session['utilisateur_id'] = utilisateur.id_utilisateur  
                return render(request, 'page_acceuil.html')
            else:
                mot_de_passe_invalid = True
        except Utilisateur.DoesNotExist:
            nom_utilisateur_invalid = True

    return render(request, 'login.html', {'nom_utilisateur_invalid': nom_utilisateur_invalid, 'mot_de_passe_invalid': mot_de_passe_invalid})



@require_http_methods(["GET", "POST"])
def index(request):
    if request.method == "POST":
        if "excel_file" in request.FILES:
            excel_file = request.FILES["excel_file"]
            try:
                df = pd.read_excel(excel_file, decimal=',', engine='openpyxl')

                
                numeric_columns = ['Stock Initial', 'Apports pour Consommation Interne', 
                                   'Prélèvement ou consommation interne', 'Prélèvements pour la Consommation autres périmètres',
                                   'Pertes', 'Expédition vers TRC', 'Livraison']

                missing_columns = [col for col in numeric_columns if col not in df.columns]
                if missing_columns:
                    return HttpResponseBadRequest(f"Colonnes manquantes dans le fichier Excel : {', '.join(missing_columns)}")

                
                for col in numeric_columns:
                    df[col] = df[col].astype(str).str.replace(r'[ .]', '', regex=True)

                df['Production'] = df['Production'].astype(str).str.replace('.', ',', regex=False)
                df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')
                df['Production'] = pd.to_numeric(df['Production'], errors='coerce')  

                test_result = True
                for index, row in df.iloc[1:-1].iterrows():
                    test = row['Stock Initial'] + row['Apports pour Consommation Interne'] + row['Production'] - \
                           row['Prélèvement ou consommation interne'] - row['Prélèvements pour la Consommation autres périmètres'] - \
                           row['Pertes'] - row['Expédition vers TRC'] - row['Livraison']
                    if abs(test) > 0.1:  
                        test_result = False
                        failed_row_index = index + 2  
                        break
                if test_result:
                    fs = FileSystemStorage()
                    filename = fs.save(excel_file.name, excel_file)
                    file_path = os.path.join(settings.MEDIA_ROOT, filename)
                    request.session['excel_file_path'] = file_path
                    return redirect('application:page_resultat_verifier')
                else:
                    return render(request, 'page_resultat_non_verifier.html', {'failed_row_index': failed_row_index})

            except pd.errors.EmptyDataError:
                return HttpResponseBadRequest("Le fichier Excel est vide.")  
            except pd.errors.ParserError:
                return HttpResponseBadRequest("Erreur lors de la lecture du fichier Excel. Vérifiez le format.")  
            except KeyError as e:
                return HttpResponseBadRequest(f"Colonne manquante dans le fichier Excel : {e}") 
            except ValueError as e:
                return HttpResponseBadRequest(str(e)) 
            except Exception as e:  
                return HttpResponseServerError(f"Une erreur inattendue s'est produite : {e}")  
        else:
            return HttpResponseBadRequest("Veuillez sélectionner un fichier.")  

    else:
        return render(request, "page_acceuil.html") 



def page_resultat_verifier(request):
    if request.method == 'POST':
        return save_data(request)  
    else:
        excel_file_path = request.session.get('excel_file_path')
        if not excel_file_path:
            return redirect('index')  
        return render(request, 'page_resultat_verifier.html', {'excel_file_path': excel_file_path})


def save_data(request):
    excel_file_path = request.session.get('excel_file_path')

    if not excel_file_path:
        return render(request, 'page_acceuil.html', {'error': "Aucun fichier sélectionné."})

    try:
        df = pd.read_excel(excel_file_path, decimal=',')
        numeric_columns = ['Stock Initial', 'Apports pour Consommation Interne',
                           'Prélèvement ou consommation interne',
                           'Prélèvements pour la Consommation autres périmètres',
                           'Pertes', 'Expédition vers TRC', 'Livraison',
                           'Stock final']
        for col in numeric_columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        df.fillna(0, inplace=True)
        mois = df.iloc[0, 13]
        annee = df.iloc[0, 14]

        perimetres_inexistants = []

        for index, row in df.iloc[1:-1].iterrows():
            perimetre_name = row['Périmètre']

            try:
                perimetre = Périmètre.objects.get(nom=perimetre_name)

                
                if Fichier_mansuelle.objects.filter(mois=mois, annee=annee, périmètre=perimetre).exists():
                                        return render(request, 'page_réponce.html', {'message': "existe déjà" , 'error': False} )
                production = (
                    row['Prélèvement ou consommation interne'] +
                    row['Prélèvements pour la Consommation autres périmètres'] +
                    row['Pertes'] +
                    row['Expédition vers TRC'] -
                    row['Livraison'] +
                    df.iloc[index]['Stock final'] - 
                    row['Stock Initial'] -
                    row['Apports pour Consommation Interne']
                )

                fichier_mansuelle = Fichier_mansuelle(
                    mois=mois,
                    annee=annee,
                    stock_ini=row['Stock Initial'],
                    Apport_consommation=row['Apports pour Consommation Interne'],
                    produit=production,
                    consomme=row['Prélèvement ou consommation interne'],
                    preleve=row['Prélèvements pour la Consommation autres périmètres'],
                    pertes=row['Pertes'],
                    expedie=row['Expédition vers TRC'],
                    laivraison=row['Livraison'],
                    densite=row['Densite'],
                    périmètre=perimetre, 
                )
                fichier_mansuelle.save()

            except Périmètre.DoesNotExist:
                perimetres_inexistants.append(perimetre_name)

        if perimetres_inexistants:
            return render(request, 'page_réponce.html', {'message': "Périmètres inexistants"}) 
        else:
            return render(request, 'page_réponce.html', {'message': "Fichier sauvegardé" ,'error': True})


    except FileNotFoundError:
        return render(request, 'page_acceuil.html', {'error': "Le fichier sélectionné n'a pas été trouvé."})



def tableau_regions(request):
    somme_attributs_par_region = {}

    if request.method == 'POST':
        mois = request.POST.get('mois')
        annee = request.POST.get('annee')

        if not mois or not annee:  
            return render(request, 'traitement_p1.html', {'mois': mois, 'annee': annee})

        regions = Region.objects.all()

        for region in regions:
            perimetres = Périmètre.objects.filter(region=region)
            fichiers_par_region = Fichier_mansuelle.objects.filter(périmètre__in=perimetres, mois=mois, annee=annee)

            somme_stock_ini = fichiers_par_region.aggregate(somme_stock_ini=Sum('stock_ini', default=0, output_field=FloatField()))['somme_stock_ini']
            somme_apport_consommation = fichiers_par_region.aggregate(somme_apport_consommation=Sum('Apport_consommation', default=0, output_field=FloatField()))['somme_apport_consommation']
            somme_produit = fichiers_par_region.aggregate(somme_produit=Sum('produit', default=0, output_field=FloatField()))['somme_produit']
            somme_consomme = fichiers_par_region.aggregate(somme_consomme=Sum('consomme', default=0, output_field=FloatField()))['somme_consomme']
            somme_preleve = fichiers_par_region.aggregate(somme_preleve=Sum('preleve', default=0, output_field=FloatField()))['somme_preleve']
            somme_pertes = fichiers_par_region.aggregate(somme_pertes=Sum('pertes', default=0, output_field=FloatField()))['somme_pertes']
            somme_expedie = fichiers_par_region.aggregate(somme_expedie=Sum('expedie', default=0, output_field=FloatField()))['somme_expedie']
            somme_laivraison = fichiers_par_region.aggregate(somme_laivraison=Sum('laivraison', default=0, output_field=FloatField()))['somme_laivraison']

            somme_exp_trc = fichiers_par_region.aggregate(somme_exp_trc=Sum(F('expedie') / F('densite'), default=0, output_field=FloatField()))['somme_exp_trc']

     
            somme_stock_final = somme_produit - somme_preleve - somme_pertes - somme_expedie + somme_laivraison + somme_stock_ini - somme_apport_consommation

            somme_stock_ini = "{:.3f}".format(somme_stock_ini or 0)
            somme_apport_consommation = "{:.3f}".format(somme_apport_consommation or 0)
            somme_produit = "{:.3f}".format(somme_produit or 0)
            somme_consomme = "{:.3f}".format(somme_consomme or 0)
            somme_preleve = "{:.3f}".format(somme_preleve or 0)
            somme_pertes = "{:.3f}".format(somme_pertes or 0)
            somme_expedie = "{:.3f}".format(somme_expedie or 0)
            somme_laivraison = "{:.3f}".format(somme_laivraison or 0)
            somme_exp_trc = "{:.3f}".format(somme_exp_trc or 0)
            somme_stock_final = "{:.3f}".format(somme_stock_final or 0)

            
            somme_attributs_par_region[region] = {
                'somme_stock_ini': somme_stock_ini,
                'somme_apport_consommation': somme_apport_consommation,
                'somme_produit': somme_produit,
                'somme_consomme': somme_consomme,
                'somme_preleve': somme_preleve,
                'somme_pertes': somme_pertes,
                'somme_expedie': somme_expedie,
                'somme_laivraison': somme_laivraison,
                'somme_exp_trc': somme_exp_trc,
                'somme_stock_final': somme_stock_final,
            }

    return render(request, 'traitement_p1.html', {'somme_attributs_par_region': somme_attributs_par_region, 'mois': mois, 'annee': annee})


def generate_pdf(request):
    if request.method == 'POST':
        mois = request.POST.get('mois')
        annee = request.POST.get('annee')
        
        if not mois or not annee:  
            return render(request, 'traitement_p1.html', {'mois': mois, 'annee': annee})
        
        somme_attributs_par_region = {}
        regions = Region.objects.all()

        for region in regions:
            perimetres = Périmètre.objects.filter(region=region)
            fichiers_par_region = Fichier_mansuelle.objects.filter(périmètre__in=perimetres, mois=mois, annee=annee)

            somme_perimetre_attributs = []
            for perimetre in perimetres:
                somme_stock_ini = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_stock_ini=Coalesce(Sum('stock_ini'), Value(0), output_field=FloatField()))['somme_stock_ini'])
                somme_apport_consommation = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_apport_consommation=Coalesce(Sum('Apport_consommation'), Value(0), output_field=FloatField()))['somme_apport_consommation'])
                somme_produit = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_produit=Coalesce(Sum('produit'), Value(0), output_field=FloatField()))['somme_produit'])
                somme_consomme = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_consomme=Coalesce(Sum('consomme'), Value(0), output_field=FloatField()))['somme_consomme'])
                somme_preleve = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_preleve=Coalesce(Sum('preleve'), Value(0), output_field=FloatField()))['somme_preleve'])
                somme_pertes = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_pertes=Coalesce(Sum('pertes'), Value(0), output_field=FloatField()))['somme_pertes'])
                somme_expedie = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_expedie=Coalesce(Sum('expedie'), Value(0), output_field=FloatField()))['somme_expedie'])
                somme_laivraison = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_laivraison=Coalesce(Sum('laivraison'), Value(0), output_field=FloatField()))['somme_laivraison'])
                somme_exp_trc = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_exp_trc=Coalesce(Sum(F('expedie') / F('densite')), Value(0), output_field=FloatField()))['somme_exp_trc'])

                somme_stock_final = "{:.3f}".format(float(somme_produit) - float(somme_preleve) - float(somme_pertes) - float(somme_expedie) + float(somme_laivraison) + float(somme_stock_ini) - float(somme_apport_consommation))

                somme_perimetre_attributs.append({
                    'nom': perimetre.nom,
                    'somme_stock_ini': somme_stock_ini,
                    'somme_apport_consommation': somme_apport_consommation,
                    'somme_produit': somme_produit,
                    'somme_consomme': somme_consomme,
                    'somme_preleve': somme_preleve,
                    'somme_pertes': somme_pertes,
                    'somme_expedie': somme_expedie,
                    'somme_laivraison': somme_laivraison,
                    'somme_exp_trc': somme_exp_trc,
                    'somme_stock_final': somme_stock_final
                })

            somme_attributs_par_region[region] = somme_perimetre_attributs

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'filename="regions_summary.pdf"'
        doc = SimpleDocTemplate(response, pagesize=landscape(A4),  
                           leftMargin=0.1*cm, rightMargin=0.1*cm, 
                           topMargin=0.1*cm, bottomMargin=0.1*cm)

        headers = ['Région', 'Périmètre', 'Stock initial', 'Consommation I', 'Production', 'Prélévement CI',
                   'Prélévement CP', 'Pertes', 'Expédition vers TRC', 'Livraison', 'Expédition TRC en m3', 'Stock final']

        col_widths = [3.4 * cm, 3.7 * cm, 2 * cm,2.3 * cm,2 * cm, 2.5 * cm,2.5 * cm, 1.5 * cm, 3.1 * cm, 1.5 * cm, 3.1 * cm, 2 * cm]
        table_data = [headers]
        for region, perimetres in somme_attributs_par_region.items():
            for perimetre_data in perimetres:
                row_data = [region.nom, perimetre_data['nom']]
                for attribut in ['somme_stock_ini', 'somme_apport_consommation', 'somme_produit', 'somme_consomme', 'somme_preleve', 'somme_pertes', 'somme_expedie', 'somme_laivraison', 'somme_exp_trc', 'somme_stock_final']:
                    row_data.append(perimetre_data[attribut])
                table_data.append(row_data)

            perimetre_totals = ['Total pour ' + region.nom, '', *[float("{:.3f}".format(sum(float(perimetre_data[attribut]) for perimetre_data in perimetres))) for attribut in ['somme_stock_ini', 'somme_apport_consommation', 'somme_produit', 'somme_consomme', 'somme_preleve', 'somme_pertes', 'somme_expedie', 'somme_laivraison', 'somme_exp_trc', 'somme_stock_final']]]
            table_data.append(perimetre_totals)

       
        table = Table(table_data, colWidths=col_widths)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.orange), 
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('FONTSIZE', (0, 0), (-1, 0), 8), 
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  
            ('GRID', (0, 0), (-1, -1), 1, colors.black)  
        ])

        table.setStyle(style)
        doc.build([table])
        return response

 
def generate_excel(request):
    if request.method == 'POST':
        mois = request.POST.get('mois')
        annee = request.POST.get('annee')
        
        if not mois or not annee: 
            return render(request, 'traitement_p1.html', {'mois': mois, 'annee': annee})
        
        regions = Region.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Régions"
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        black_border = Border(left=Side(style='thin', color='000000'),
                              right=Side(style='thin', color='000000'),
                              top=Side(style='thin', color='000000'),
                              bottom=Side(style='thin', color='000000'))

        ws['A1'] = 'Mois : ' + mois
        ws['A2'] = 'Année : ' + annee
        ws.merge_cells('A1:B1')
        ws.merge_cells('A2:B2')

        ws['A1'].font = bold_font
        ws['A1'].alignment = center_alignment
        ws['A2'].font = bold_font
        ws['A2'].alignment = center_alignment

        
        headers = ['Région', 'Périmètre', 'Stock initial', 'Apport de consommation interne', 'Production', 'Consommation interne',
                   'Prélévement ou consommation autre périmétre', 'Pertes', 'Expédition vers TRC', 'Livraison', 'Expédition vers TRC en m3', 'Stock final']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.fill = orange_fill

        row_num = 4  

        for region in regions:
            perimetres = Périmètre.objects.filter(region=region)
            fichiers_par_region = Fichier_mansuelle.objects.filter(périmètre__in=perimetres, mois=mois, annee=annee)
            for perimetre in perimetres:
                attributs = [
                    region.nom,
                    perimetre.nom,
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_stock_ini=Coalesce(Sum('stock_ini'), Value(0), output_field=FloatField()))['somme_stock_ini'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_apport_consommation=Coalesce(Sum('Apport_consommation'), Value(0), output_field=FloatField()))['somme_apport_consommation'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_produit=Coalesce(Sum('produit'), Value(0), output_field=FloatField()))['somme_produit'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_consomme=Coalesce(Sum('consomme'), Value(0), output_field=FloatField()))['somme_consomme'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_preleve=Coalesce(Sum('preleve'), Value(0), output_field=FloatField()))['somme_preleve'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_pertes=Coalesce(Sum('pertes'), Value(0), output_field=FloatField()))['somme_pertes'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_expedie=Coalesce(Sum('expedie'), Value(0), output_field=FloatField()))['somme_expedie'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_laivraison=Coalesce(Sum('laivraison'), Value(0), output_field=FloatField()))['somme_laivraison'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_exp_trc=Coalesce(Sum(F('expedie') / F('densite')), Value(0), output_field=FloatField()))['somme_exp_trc'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_stock_final=Coalesce(Sum(F('produit') - F('preleve') - F('pertes') - F('expedie') + F('laivraison') + F('stock_ini') - F('Apport_consommation')), Value(0), output_field=FloatField()))['somme_stock_final']
                ]

                
                for col_num, value in enumerate(attributs, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = center_alignment

                row_num += 1  

            total_attributs = []
            for col_num in range(3, ws.max_column + 1):
                total = sum(ws.cell(row=row, column=col_num).value or 0 for row in range(row_num - len(perimetres), row_num))
                total_attributs.append(total)

            for col_num, value in enumerate(total_attributs, 3):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = bold_font
                cell.alignment = center_alignment

            cell = ws.cell(row=row_num, column=1, value="Total pour " + region.nom)
            cell.font = bold_font
            cell.alignment = center_alignment

            row_num += 1  
        for col_num in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].auto_size = True

        for row in ws.iter_rows():
            for cell in row:
                cell.border = black_border
        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)

        response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=bilan_par_périmétre.xlsx'
        return response

def tableau_regions_annuel(request, annee=None):
    somme_attributs_par_region = {}
    
    if request.method == 'POST':
        annee = request.POST.get('annee')
    if not annee:  
            return render(request, 'traitement_annuel.html', {'annee': annee})
    if annee:
        regions = Region.objects.all()

        for region in regions:
            perimetres = Périmètre.objects.filter(region=region)
            fichiers_par_region = Fichier_mansuelle.objects.filter(périmètre__in=perimetres, annee=annee)
        
            somme_stock_ini = fichiers_par_region.aggregate(somme_stock_ini=Sum('stock_ini', default=0, output_field=FloatField()))['somme_stock_ini']
            somme_apport_consommation = fichiers_par_region.aggregate(somme_apport_consommation=Sum('Apport_consommation', default=0, output_field=FloatField()))['somme_apport_consommation']
            somme_produit = fichiers_par_region.aggregate(somme_produit=Sum('produit', default=0, output_field=FloatField()))['somme_produit']
            somme_consomme = fichiers_par_region.aggregate(somme_consomme=Sum('consomme', default=0, output_field=FloatField()))['somme_consomme']
            somme_preleve = fichiers_par_region.aggregate(somme_preleve=Sum('preleve', default=0, output_field=FloatField()))['somme_preleve']
            somme_pertes = fichiers_par_region.aggregate(somme_pertes=Sum('pertes', default=0, output_field=FloatField()))['somme_pertes']
            somme_expedie = fichiers_par_region.aggregate(somme_expedie=Sum('expedie', default=0, output_field=FloatField()))['somme_expedie']
            somme_laivraison = fichiers_par_region.aggregate(somme_laivraison=Sum('laivraison', default=0, output_field=FloatField()))['somme_laivraison']

            somme_exp_trc = fichiers_par_region.aggregate(somme_exp_trc=Sum(F('expedie') / F('densite'), default=0, output_field=FloatField()))['somme_exp_trc']

            somme_stock_final = somme_produit - somme_preleve - somme_pertes - somme_expedie + somme_laivraison + somme_stock_ini - somme_apport_consommation
            
            somme_stock_ini = "{:.3f}".format(somme_stock_ini or 0)
            somme_apport_consommation = "{:.3f}".format(somme_apport_consommation or 0)
            somme_produit = "{:.3f}".format(somme_produit or 0)
            somme_consomme = "{:.3f}".format(somme_consomme or 0)
            somme_preleve = "{:.3f}".format(somme_preleve or 0)
            somme_pertes = "{:.3f}".format(somme_pertes or 0)
            somme_expedie = "{:.3f}".format(somme_expedie or 0)
            somme_laivraison = "{:.3f}".format(somme_laivraison or 0)
            somme_exp_trc = "{:.3f}".format(somme_exp_trc or 0)
            somme_stock_final = "{:.3f}".format(somme_stock_final or 0)
        
            somme_attributs_par_region[region] = {
                'somme_stock_ini': somme_stock_ini,
                'somme_apport_consommation': somme_apport_consommation,
                'somme_produit': somme_produit,
                'somme_consomme': somme_consomme,
                'somme_preleve': somme_preleve,
                'somme_pertes': somme_pertes,
                'somme_expedie': somme_expedie,
                'somme_laivraison': somme_laivraison,
                'somme_exp_trc': somme_exp_trc,
                'somme_stock_final': somme_stock_final,
            }

    return render(request, 'traitement_annuel.html', {'somme_attributs_par_region': somme_attributs_par_region, 'annee': annee})



def generate_pdf_annuel(request):
    if request.method == 'POST':
        annee = request.POST.get('annee')

        if not annee: 
            return render(request, 'traitement_annuel.html', {'annee': annee})

        somme_attributs_par_region = {}
        regions = Region.objects.all()

        for region in regions:
            perimetres = Périmètre.objects.filter(region=region)
            fichiers_par_region = Fichier_mansuelle.objects.filter(périmètre__in=perimetres, annee=annee)
            
            somme_perimetre_attributs = []
            for perimetre in perimetres:
                somme_stock_ini = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_stock_ini=Coalesce(Sum('stock_ini'), Value(0), output_field=FloatField()))['somme_stock_ini'])
                somme_apport_consommation = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_apport_consommation=Coalesce(Sum('Apport_consommation'), Value(0), output_field=FloatField()))['somme_apport_consommation'])
                somme_produit = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_produit=Coalesce(Sum('produit'), Value(0), output_field=FloatField()))['somme_produit'])
                somme_consomme = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_consomme=Coalesce(Sum('consomme'), Value(0), output_field=FloatField()))['somme_consomme'])
                somme_preleve = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_preleve=Coalesce(Sum('preleve'), Value(0), output_field=FloatField()))['somme_preleve'])
                somme_pertes = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_pertes=Coalesce(Sum('pertes'), Value(0), output_field=FloatField()))['somme_pertes'])
                somme_expedie = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_expedie=Coalesce(Sum('expedie'), Value(0), output_field=FloatField()))['somme_expedie'])
                somme_laivraison = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_laivraison=Coalesce(Sum('laivraison'), Value(0), output_field=FloatField()))['somme_laivraison'])
                somme_exp_trc = "{:.3f}".format(fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_exp_trc=Coalesce(Sum(F('expedie') / F('densite')), Value(0), output_field=FloatField()))['somme_exp_trc'])

                somme_stock_final = "{:.3f}".format(float(somme_produit) - float(somme_preleve) - float(somme_pertes) - float(somme_expedie) + float(somme_laivraison) + float(somme_stock_ini) - float(somme_apport_consommation))

                somme_perimetre_attributs.append({
                    'nom': perimetre.nom,
                    'somme_stock_ini': somme_stock_ini,
                    'somme_apport_consommation': somme_apport_consommation,
                    'somme_produit': somme_produit,
                    'somme_consomme': somme_consomme,
                    'somme_preleve': somme_preleve,
                    'somme_pertes': somme_pertes,
                    'somme_expedie': somme_expedie,
                    'somme_laivraison': somme_laivraison,
                    'somme_exp_trc': somme_exp_trc,
                    'somme_stock_final': somme_stock_final
                })

            somme_attributs_par_region[region] = somme_perimetre_attributs

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'filename="regions_summary.pdf"'
        doc = SimpleDocTemplate(response, pagesize=landscape(A4),  
                           leftMargin=0.1*cm, rightMargin=0.1*cm, 
                           topMargin=0.1*cm, bottomMargin=0.1*cm)
        
        headers = ['Région', 'Périmètre', 'Stock initial', 'Consommation I', 'Production', 'Prélévement CI',
                   'Prélévement CP', 'Pertes', 'Expédition vers TRC', 'Livraison', 'Expédition TRC en m3', 'Stock final']

        col_widths = [3.4 * cm, 3.7 * cm, 2 * cm,2.3 * cm,2 * cm, 2.5 * cm,2.5 * cm, 1.5 * cm, 3.1 * cm, 1.5 * cm, 3.1 * cm, 2 * cm]
        table_data = [headers]
        for region, perimetres in somme_attributs_par_region.items():
            for perimetre_data in perimetres:
                row_data = [region.nom, perimetre_data['nom']]
                for attribut in ['somme_stock_ini', 'somme_apport_consommation', 'somme_produit', 'somme_consomme', 'somme_preleve', 'somme_pertes', 'somme_expedie', 'somme_laivraison', 'somme_exp_trc', 'somme_stock_final']:
                    row_data.append(perimetre_data[attribut])
                table_data.append(row_data)

            perimetre_totals = ['Total pour ' + region.nom, '', *[float("{:.3f}".format(sum(float(perimetre_data[attribut]) for perimetre_data in perimetres))) for attribut in ['somme_stock_ini', 'somme_apport_consommation', 'somme_produit', 'somme_consomme', 'somme_preleve', 'somme_pertes', 'somme_expedie', 'somme_laivraison', 'somme_exp_trc', 'somme_stock_final']]]
            table_data.append(perimetre_totals)

        table = Table(table_data, colWidths=col_widths)

      
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.orange), 
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black), 
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12), 
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('FONTSIZE', (0, 0), (-1, 0), 8), 
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige), 
            ('GRID', (0, 0), (-1, -1), 1, colors.black)  
        ])

        table.setStyle(style)
        doc.build([table])
        return response


def generate_excel_annuele(request):
    if request.method == 'POST':
        annee = request.POST.get('annee')

        if not annee:  
            return render(request, 'traitement_annuel.html', {'annee': annee})

        regions = Region.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Régions"
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        black_border = Border(left=Side(style='thin', color='000000'),
                              right=Side(style='thin', color='000000'),
                              top=Side(style='thin', color='000000'),
                              bottom=Side(style='thin', color='000000'))

        
        ws['A1'] = 'Année : ' + annee

        ws.merge_cells('A1:B1')
        ws['A1'].font = bold_font
        ws['A1'].alignment = center_alignment
        headers = ['Région', 'Périmètre', 'Stock initial', 'Apport de consommation interne', 'Production', 'Consommation',
                   'Prélevé', 'Pertes', 'Expédition vers TRC', 'Livraison', 'Expédition vers TRC en m3', 'Stock final']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.fill = orange_fill

        row_num = 4  
        for region in regions:
            perimetres = Périmètre.objects.filter(region=region)
            fichiers_par_region = Fichier_mansuelle.objects.filter(périmètre__in=perimetres, annee=annee)

            for perimetre in perimetres:
                attributs = [
                    region.nom,
                    perimetre.nom,
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_stock_ini=Coalesce(Sum('stock_ini'), Value(0), output_field=FloatField()))['somme_stock_ini'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_apport_consommation=Coalesce(Sum('Apport_consommation'), Value(0), output_field=FloatField()))['somme_apport_consommation'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_produit=Coalesce(Sum('produit'), Value(0), output_field=FloatField()))['somme_produit'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_consomme=Coalesce(Sum('consomme'), Value(0), output_field=FloatField()))['somme_consomme'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_preleve=Coalesce(Sum('preleve'), Value(0), output_field=FloatField()))['somme_preleve'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_pertes=Coalesce(Sum('pertes'), Value(0), output_field=FloatField()))['somme_pertes'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_expedie=Coalesce(Sum('expedie'), Value(0), output_field=FloatField()))['somme_expedie'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_laivraison=Coalesce(Sum('laivraison'), Value(0), output_field=FloatField()))['somme_laivraison'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_exp_trc=Coalesce(Sum(F('expedie') / F('densite')), Value(0), output_field=FloatField()))['somme_exp_trc'],
                    fichiers_par_region.filter(périmètre=perimetre).aggregate(somme_stock_final=Coalesce(Sum(F('produit') - F('preleve') - F('pertes') - F('expedie') + F('laivraison') + F('stock_ini') - F('Apport_consommation')), Value(0), output_field=FloatField()))['somme_stock_final']
                ]

                for col_num, value in enumerate(attributs, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = center_alignment

                row_num += 1 
            total_attributs = []
            for col_num in range(3, ws.max_column + 1):
                total = sum(ws.cell(row=row, column=col_num).value or 0 for row in range(row_num - len(perimetres), row_num))
                total_attributs.append(total)

            for col_num, value in enumerate(total_attributs, 3):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = bold_font
                cell.alignment = center_alignment
            cell = ws.cell(row=row_num, column=1, value="Total pour " + region.nom)
            cell.font = bold_font
            cell.alignment = center_alignment

            row_num += 1  
        for col_num in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].auto_size = True

        for row in ws.iter_rows():
            for cell in row:
                cell.border = black_border
        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=bilan_annuel.xlsx'

        return response


logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
handler = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)
from datetime import datetime
from django.db.models import Sum, Count, Q
from calendar import month_abbr
from .models import Fichier_mansuelle, Prévision_perimetre

def extract_data_for_visualization(date_debut: str, date_fin: str, region: str) -> dict:
    try:
        date_format = "%Y-%m"
        date_debut = datetime.strptime(date_debut, date_format)
        date_fin = datetime.strptime(date_fin, date_format)
        fichiers = Fichier_mansuelle.objects.filter(
            Q(annee__gt=date_debut.year) | 
            (Q(annee=date_debut.year) & Q(mois__gte=date_debut.month)),
            Q(annee__lt=date_fin.year) | 
            (Q(annee=date_fin.year) & Q(mois__lte=date_fin.month)),
            Q(périmètre__region__nom=region)
        )

        total_production = fichiers.aggregate(total=Sum('produit'))['total'] or 0
        production_par_perimetre = fichiers.values('périmètre__nom').annotate(production=Sum('produit'))
        data_pie_chart = [
            {
                'perimetre': item['périmètre__nom'],
                'percentage': (item['production'] / total_production) * 100 if total_production != 0 else 0
            }
            for item in production_par_perimetre
        ]
        for data in data_pie_chart:
            data['percentage'] = float(data['percentage'])

        moyenne_production_mensuelle = fichiers.values('mois', 'périmètre__nom').annotate(
            moyenne_production=Sum('produit') / Count('mois')
        )

        previsions_par_perimetre_mois = (
            Prévision_perimetre.objects
            .filter(
                Q(annee__gt=date_debut.year) | 
                (Q(annee=date_debut.year) & Q(mois__gte=date_debut.month)),
                Q(annee__lt=date_fin.year) | 
                (Q(annee=date_fin.year) & Q(mois__lte=date_fin.month)),
                Q(périmètre__region__nom=region)
            )
            .values('mois', 'périmètre__nom')
            .annotate(prevision=Sum('prévision'))
            .order_by('mois', 'périmètre__nom')
        )
        production_previsions = []
        for item in moyenne_production_mensuelle:
            prevision_item = previsions_par_perimetre_mois.filter(
                mois=item['mois'], périmètre__nom=item['périmètre__nom']
            ).first()
            production_previsions.append({
                'perimetre': item['périmètre__nom'],
                'production_mensuelle': float(item['moyenne_production']),
                'prevision': float(prevision_item['prevision']) if prevision_item else 0.0,
            })
        production_previsions_mois = [
            {
                'mois': item['mois'],
                'production_mensuelle': float(item['moyenne_production']),
                'prevision': previsions_par_perimetre_mois.filter(mois=item['mois']).aggregate(prevision=Sum('prévision'))['prevision'] or 0
            }
            for item in moyenne_production_mensuelle
        ]
        for d in production_previsions_mois:
            d['production_mensuelle'] = float(d['production_mensuelle'])
            d['prevision'] = float(d['prevision'])

        return {
            'data_pie_chart': data_pie_chart,
            'total': total_production,
            'production_previsions': production_previsions,
            'production_previsions_mois': production_previsions_mois,
        }

    except IntegrityError as ie:
        logger.warning(f"IntegrityError occurred: {ie}")
        pass  
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        raise

def get_chart_data(request):
    if request.method == 'POST':
        date_debut = request.POST.get('dateDebut')
        date_fin = request.POST.get('dateFin')
        region_name = request.POST.get('region')
    else:  
        date_debut = "2023-11"
        date_fin = "2023-12"
        region_name = "Adrar"  
    data = extract_data_for_visualization(date_debut, date_fin, region_name)
    user_id = request.session.get('utilisateur_id')
    if user_id:
        try:
            utilisateur = Utilisateur.objects.get(id_utilisateur=user_id)
            region_obj = Region.objects.get(nom=region_name)
            Visualisation.objects.create(
                date=timezone.now(),
                date_debut=datetime.strptime(date_debut, "%Y-%m").date(),
                date_fin=datetime.strptime(date_fin, "%Y-%m").date(),
                region=region_obj,
                utilisateur=utilisateur,
            )

        except (Utilisateur.DoesNotExist, Region.DoesNotExist, ValueError):
            pass  
    
    context = {
        "data_pie_chart": data['data_pie_chart'],
        "total": data['total'],
        'productionPrev': data['production_previsions'],
        'production_previsions_mois': data['production_previsions_mois'],
        'dateDeb': date_debut,
        'dateFin': date_fin,
        'region': region_name,  
    }

    return render(request, 'dashboard.html', context=context)


def generate_excel_from_extract(request):
    if request.method == 'POST':
        date_debut_str = request.POST.get('dateDebut')
        date_fin_str = request.POST.get('dateFin')
        region_nom = request.POST.get('region')
        error_message = request.POST.get('error', '')

        try:
            date_debut = datetime.strptime(date_debut_str, "%Y-%m")
            date_fin = datetime.strptime(date_fin_str, "%Y-%m")
            region = Region.objects.get(nom=region_nom)
        except ValueError:
            error_message = "Format de date incorrect. Utilisez YYYY-MM."
        except Region.DoesNotExist:
            error_message = "Région introuvable."

        if error_message:
            request.POST._mutable = True
            request.POST['error'] = error_message
            return render(request, 'dashboard.html', request.POST)

        mois_francais = {
            'janvier': '01', 'février': '02', 'mars': '03', 'avril': '04',
            'mai': '05', 'juin': '06', 'juillet': '07', 'août': '08',
            'septembre': '09', 'octobre': '10', 'novembre': '11', 'décembre': '12'
        }

        try:
            Fichier_mansuelle.objects.filter(mois__in=mois_francais.keys()).update(
                mois=Case(
                    *[When(mois=k, then=Value(v)) for k, v in mois_francais.items()],
                    default=F('mois')
                )
            )
        except IntegrityError:
            pass

        fichiers = Fichier_mansuelle.objects.filter(
            Q(annee__gte=date_debut.year, mois__gte=date_debut.month) &
            Q(annee__lte=date_fin.year, mois__lte=date_fin.month) &
            Q(périmètre__region=region)
        )

        production_par_perimetre = fichiers.values('périmètre__nom').annotate(production=Sum('produit'))
        total_production = fichiers.aggregate(total=Sum('produit'))['total'] or 0

        moyenne_production_mensuelle = fichiers.values('mois', 'périmètre__nom').annotate(moyenne_production=Sum('produit') / Count('périmètre'))

        previsions = Prévision_perimetre.objects.filter(
            Q(annee__gte=date_debut.year, mois__gte=date_debut.month) &
            Q(annee__lte=date_fin.year, mois__lte=date_fin.month) &
            Q(périmètre__region=region)
        )

        production_previsions = [
            {
                'perimetre': item['périmètre__nom'],
                'production_mensuelle': item['moyenne_production'],
                'prevision': previsions.filter(mois=item['mois']).aggregate(prevision=Sum('prévision'))['prevision'] or 0
            }
            for item in moyenne_production_mensuelle
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = f"Données de {region_nom}"

        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:C1')
        ws['A1'] = f"Région: {region_nom} - Période: {date_debut_str} à {date_fin_str}"
        ws['A1'].font = bold_font
        ws['A1'].alignment = center_alignment

        headers = ['Périmètre', 'Production', 'Prévision']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)  
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.fill = orange_fill
            cell.border = thin_border

      
        row_num = 4 
        for data in production_previsions:
            ws.cell(row=row_num, column=1, value=data['perimetre']).border = thin_border
            ws.cell(row=row_num, column=2, value=data['production_mensuelle']).border = thin_border
            ws.cell(row=row_num, column=3, value=data['prevision']).border = thin_border
            row_num += 1
        ws.cell(row=row_num, column=1, value="Total").font = bold_font
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=2, value=total_production).font = bold_font
        ws.cell(row=row_num, column=2).border = thin_border

        for col_idx in range(1, ws.max_column + 1):  
           max_length = 0
           column_letter = get_column_letter(col_idx)  
           for row_idx in range(1, ws.max_row + 1):  
             cell = ws.cell(row=row_idx, column=col_idx)
             if cell.value:
                max_length = max(max_length, len(str(cell.value)))
    
             adjusted_width = max_length + 2  
             ws.column_dimensions[column_letter].width = adjusted_width

      
        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)

        response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=donnees_{region_nom}_{date_debut_str}_{date_fin_str}.xlsx'

        return response

